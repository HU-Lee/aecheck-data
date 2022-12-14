import os
import re
import shutil
from openpyxl import load_workbook
from config import FILE_NAME, ELEMENTS, WEAPONS, IMG_ROOT, IMG_TAG
from function import make_folder, write_json


workbook = load_workbook(FILE_NAME, data_only=True)
char_sheet = workbook["캐릭터"]
buddy_sheet = workbook["버디"]
dun_sheet = workbook["던전"]
per_sheet = workbook["퍼스널리티"]


for name in ["character", "buddy", "personality", "result_json", "result_json/language"]:
    make_folder(name)


"""
캐릭터 이미지 저장
"""
for row in char_sheet.rows:
    filename = "{}{}_command.png".format(row[3].value, IMG_TAG.get(str(row[2].value), "jsalkjdakljdlajdalkd"))
    
    # 테일즈 캐릭터는 5성 이미지만 있어서 보정
    if row[1].value in ["클레스", "유리", "미라", "벨벳"] and str(row[2].value)=="4.5":
        filename = filename.replace("command", "rank5_command")
    
    # 아이디대로 이미지를 복사
    for file in os.listdir(IMG_ROOT):
        if filename in file:
            shutil.copyfile("rawimage/" + file, "character/{}.png".format(row[0].value))


"""
버디 이미지 저장
"""
for row in buddy_sheet.rows:
    filename = "{}.png".format(row[2].value)
    # 아이디대로 이미지를 복사
    for file in os.listdir(IMG_ROOT):
        if filename in file:
            shutil.copyfile("rawimage/" + file, "buddy/{}.png".format(row[0].value))


"""
퍼스널리티 이미지 저장
"""
for row in per_sheet.rows:    
    # 이름으로 이미지를 복사
    for file in os.listdir(IMG_ROOT):
        if str(row[4].value) in file and ".png" in file:
            if row[1].value and "s3" in file:
                shutil.copyfile("rawimage/" + file, "personality/{}(ES).png".format(row[0].value))
                break
            elif not row[1].value and "s3" not in file:
                shutil.copyfile("rawimage/" + file, "personality/{}.png".format(row[0].value))


"""
character.json
"""
char_arr = []
codes = []

for row in char_sheet.iter_rows(min_row=2):
    dic = {
        "id": row[0].value,
        "code": str(row[3].value),
        "style": str(row[2].value).lower(),
        "category": ELEMENTS.index(row[4].value)*10 + WEAPONS.index(row[5].value),
        "free": row[6].value,
        "sky": "light" if row[7].value == "천" else "shadow",
        "first": row[3].value not in codes,
        "jonly": row[8].value,
        "gonly": row[9].value,
        "from": list(map(int, str(row[10].value).split(","))) if row[10].value else [],
        "change": list(map(int, str(row[11].value).split(","))) if row[11].value else [],
        "book": row[12].value or "없음",
        "book_get": row[13].value.split(",") if row[13].value else [],
        "manifest_jap": row[15].value or "없음",
        "manifest_glo": row[16].value or "없음"
    }
    char_arr.append(dic)
    codes = list(set(codes + [row[3].value]))

write_json('result_json/character.json', char_arr)


"""
buddy.json
"""
buddy_arr = []

for row in buddy_sheet.iter_rows(min_row=2):
    if not row[0].value:  continue
    dic = {
        "id": row[0].value,
        "code": str(row[2].value),
        "from": list(map(int, str(row[3].value).split(","))) if row[3].value else [],
        "free": row[4].value,
        "jonly": row[5].value,
        "gonly": row[6].value,
    }
    buddy_arr.append(dic)

write_json('result_json/buddy.json', buddy_arr)


"""
번역 json
"""
kor_json = {}
eng_json = {}
jap_json = {}

# 속성, 무기 번역
etc_trans_sheet = workbook["기타번역"]
for row in etc_trans_sheet.iter_rows(min_row=2):
    eng_key = re.sub(r"[^a-zA-Z0-9]","",row[1].value).lower()
    kor_json[eng_key] = row[0].value
    eng_json[eng_key] = row[1].value
    jap_json[eng_key] = row[2].value

# 설명 번역
desc_trans_sheet = workbook["설명"]
for row in desc_trans_sheet.iter_rows(min_row=2):
    eng_key = row[3].value
    if not eng_key: 
        continue
    kor_json[eng_key] = row[0].value
    eng_json[eng_key] = row[1].value
    jap_json[eng_key] = row[2].value

# 캐릭터 이름 번역
for row in char_sheet.iter_rows(min_row=2):
    kor_json[str(row[3].value)] = row[1].value

char_trans_sheet = workbook["캐릭번역"]
for row in char_trans_sheet.iter_rows(min_row=2):
    for key in kor_json.keys():
        if kor_json[key] == row[0].value:
            eng_json[key] = row[1].value
            jap_json[key] = row[2].value

# 버디 이름 번역
for row in buddy_sheet.iter_rows(min_row=2):
    if not row[0].value: continue
    kor_json[str(row[2].value)] = row[1].value

char_trans_sheet = workbook["캐릭번역"]
for row in char_trans_sheet.iter_rows(min_row=2):
    for key in kor_json.keys():
        if kor_json[key] == row[0].value:
            eng_json[key] = row[1].value
            jap_json[key] = row[2].value

# 나머지 작업
for i in ["던전", "기타번역", "캐릭번역", "특성번역"]:
    sheet = workbook[i]
    for row in sheet.iter_rows(min_row=2):
        kor_key = row[0].value
        if kor_key == "null" or kor_key == None:
            continue
        kor_json[kor_key] = row[0].value
        eng_json[kor_key] = row[1].value
        jap_json[kor_key] = row[2].value

write_json('result_json/language/ko.json', kor_json)
write_json('result_json/language/en.json', eng_json)
write_json('result_json/language/jp.json', jap_json)


"""
dungeon.json
"""
dun_arr = list(map(lambda row: {
    "name": row[0].value,
    "endpoint": row[3].value
}, 
dun_sheet.iter_rows(min_row=2)))

write_json('result_json/dungeon.json', dun_arr)


"""
personality.json
"""
per_arr = list(map(lambda row: {
    "name": row[0].value,
    "is_extra": row[1].value,
    "personality": row[2].value.split(","),
    "description": row[3].value,
    "code": str(row[4].value)
}, 
per_sheet.iter_rows(min_row=2)))

write_json('result_json/personality.json', per_arr)